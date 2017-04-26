from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image
import sys

imgfname = sys.argv[1]

# make image thumbnail
basewidth = 300
img = Image.open(imgfname).convert('RGB')
wpercent = (basewidth/float(img.size[0]))
hsize = int((float(img.size[1])*float(wpercent)))
im = img.resize((basewidth,hsize), Image.ANTIALIAS)

width,height = im.size

# create new excel sheet
wb = Workbook()
dest_filename = ".".join(imgfname.split(".")[:-1]+["xlsx"])
print dest_filename

ws1 = wb.active
ws1.title = imgfname

print "Preparing excel file.."
for i in range(width):
	for j in range(height):
		r,g,b=im.getpixel((i,j))
		hexrgb = 'FF{:02x}{:02x}{:02x}'.format(r, g, b)
		fill = PatternFill(start_color=hexrgb,end_color=hexrgb,fill_type='solid')
		cel = ws1.cell(row=j+1, column=i+1)
		cel.fill = fill
	ws1.column_dimensions[get_column_letter(i+1)].width = 3


wb.save(filename = dest_filename)
print "Done!"